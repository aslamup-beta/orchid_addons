# -*- coding: utf-8 -*-
import io
import re
import urllib

try:
    text_type = unicode  # noqa: F821 (Python 2)
except NameError:
    text_type = str  # Python 3 fallback, just in case

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from openerp import http
from openerp.http import request


class SlaReportController(http.Controller):
    """Streams the SLA report as a generated .xlsx file so that clicking
    'Generate Report' on the wizard triggers a direct browser download.
    """

    @http.route('/bit_outsource_emp_expiry/xlsx/<int:wizard_id>', type='http',
                auth='user')
    def download_sla_report_xlsx(self, wizard_id, **kwargs):
        wizard = request.env['sla.report.wizard'].browse(wizard_id)
        if not wizard.exists():
            return request.not_found()

        data = wizard.get_report_data()
        xlsx_bytes = self._build_workbook(data)

        organisation_name = data['organisation_name'] or 'report'

        # Build a pure-ASCII fallback filename for the legacy
        # `filename=` parameter (required because some HTTP header
        # encoders, e.g. werkzeug's latin-1 codec, choke on non-ASCII
        # bytes), plus a UTF-8 `filename*=` parameter (RFC 5987) so
        # browsers display the original, non-ASCII organisation name
        # correctly.
        ascii_name = organisation_name.encode('ascii', 'ignore') \
            if isinstance(organisation_name, text_type) \
            else organisation_name
        ascii_name = re.sub(r'[^A-Za-z0-9]+', '_', ascii_name).strip('_')
        ascii_name = ascii_name or 'report'
        ascii_filename = 'SLA_Report_%s.xlsx' % ascii_name

        if isinstance(organisation_name, bytes):
            organisation_name_u = organisation_name.decode('utf-8')
        else:
            organisation_name_u = organisation_name
        utf8_filename = u'SLA_Report_%s.xlsx' % organisation_name_u
        quoted_utf8_filename = urllib.quote(
            utf8_filename.encode('utf-8')
        )

        content_disposition = (
            'attachment; filename="%s"; filename*=UTF-8\'\'%s'
            % (ascii_filename, quoted_utf8_filename)
        )

        headers = [
            ('Content-Type',
             'application/vnd.openxmlformats-officedocument'
             '.spreadsheetml.sheet'),
            ('Content-Disposition', content_disposition),
            ('Content-Length', len(xlsx_bytes)),
        ]
        return request.make_response(xlsx_bytes, headers=headers)

    def _build_workbook(self, data):
        """Build the SLA report workbook (in memory) and return its
        bytes content.

        :param data: dict as returned by
                      sla.report.wizard.get_report_data()
        :return: bytes content of the .xlsx file
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'SLA Report'

        thin = Side(style='thin', color='000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        title_font = Font(bold=True, size=16)
        label_font = Font(bold=True, size=10)
        header_font = Font(bold=True, size=10, color='FFFFFF')
        header_fill = PatternFill(
            fill_type='solid', start_color='4472C4', end_color='4472C4'
        )
        center = Alignment(horizontal='center', vertical='center')
        left = Alignment(horizontal='left', vertical='center')

        # ---- Title ----
        ws.merge_cells('A1:E1')
        ws['A1'] = 'SLA Performance Report'
        ws['A1'].font = title_font
        ws['A1'].alignment = center

        # ---- Summary section ----
        ws['A3'] = 'Customer Name:'
        ws['A3'].font = label_font
        ws['B3'] = data['organisation_name']

        ws['A4'] = 'Period:'
        ws['A4'].font = label_font
        ws['B4'] = '%s to %s' % (data['date_start'], data['date_end'])

        ws['A5'] = 'Number of tickets:'
        ws['A5'].font = label_font
        ws['B5'] = data['total_count']

        ws['A6'] = 'Number of closed tickets:'
        ws['A6'].font = label_font
        ws['B6'] = data['closed_count']

        # ---- Tickets table ----
        header_row = 8
        headers = [
            'Sequence', 'Description', 'Project', 'Status', 'Initiation Date',
            'Closing Date', 'Hour Difference', 'Resolution Time','Priority',
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        row_idx = header_row + 1
        if data['ticket_lines']:
            for line in data['ticket_lines']:
                values = [
                    line['sequence'],
                    line['description'],
                    line['project'],
                    line['status'],
                    line['initiation_date'],
                    line['closing_date'],
                    line['hours_difference'],
                    line['defined_response'],
                    line['priority'],
                ]
                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(
                        row=row_idx, column=col_idx, value=value
                    )
                    cell.border = border
                    cell.alignment = left
                row_idx += 1
        else:
            ws.merge_cells(
                start_row=row_idx, start_column=1,
                end_row=row_idx, end_column=5
            )
            no_data_cell = ws.cell(
                row=row_idx, column=1,
                value='No tickets found for the selected organisation '
                      'and date range.'
            )
            no_data_cell.alignment = center

        # ---- Column widths ----
        widths = [21,45, 45, 13, 20, 20, 15, 15, 11]
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
