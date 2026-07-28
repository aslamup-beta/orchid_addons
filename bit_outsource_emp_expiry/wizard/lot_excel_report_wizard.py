# -*- coding: utf-8 -*-
import base64
from io import BytesIO
from datetime import datetime

from openerp import models, fields, api
from openerp.exceptions import Warning as UserError

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None


class LotExcelReportWizard(models.TransientModel):
    _name = 'lot.excel.report.wizard'
    _description = 'Serial Number Excel Report Wizard'

    start_date = fields.Date(
        string='Start Date',
        required=True,
        help='Only serial numbers whose expiry date falls on or after '
             'this date will be reported.'
    )
    end_date = fields.Date(
        string='End Date',
        required=True,
        help='Only serial numbers whose expiry date falls on or before '
             'this date will be reported.'
    )
    product_id = fields.Many2one(
        'product.product',
        string='Product',
        help='Only serial numbers linked to this product will be reported. '
             'Leave empty to include all products.'
    )
    partner_id = fields.Many2one(
        'res.partner',
        string='Customer',
        help='Only serial numbers linked to this customer will be reported. '
             'Leave empty to include all customers.'
    )

    @api.constrains('start_date', 'end_date')
    def _check_date_range(self):
        for wizard in self:
            if wizard.start_date and wizard.end_date \
                    and wizard.start_date > wizard.end_date:
                raise UserError(
                    'The start date must be earlier than or equal to the '
                    'end date.'
                )

    @api.multi
    def action_generate_report(self):
        """Fetch matching stock.production.lot records and stream them
        into an .xlsx workbook using openpyxl, then return an action that
        downloads the generated file."""
        self.ensure_one()

        if openpyxl is None:
            raise UserError(
                'The "openpyxl" python library is not installed on the '
                'server. Please install it (pip install openpyxl) and '
                'restart Odoo.'
            )

        lot_obj = self.env['stock.production.lot']
        # 'life_date' holds the expiry date of the lot (provided by the
        # 'product_expiry' module). Use a half-open-day range on the
        # datetime string so that lots with a date+time value are still
        # matched correctly against the end date.
        domain = [
            ('life_date', '>=', '%s 00:00:00' % self.start_date),
            ('life_date', '<=', '%s 23:59:59' % self.end_date),
        ]
        if self.product_id:
            domain += [('product_id', '=', self.product_id.id)]
        if self.partner_id:
            domain += [('customer_id', '=', self.partner_id.id)]

        lots = lot_obj.search(domain, order='name asc')

        if not lots:
            raise UserError(
                'No serial numbers were found that expire between %s and '
                '%s for the selected filters.'
                % (self.start_date, self.end_date)
            )

        attachment = self._build_xlsx_attachment(lots)

        # '/web/binary/saveas' is the classic Odoo 8 download endpoint: it
        # forces a native "Save As" dialog based on the attachment's
        # 'datas' field and 'datas_fname', which is more reliable across
        # browsers than '/web/content' for triggering an actual download
        # (rather than opening the file inline).
        return {
            'type': 'ir.actions.act_url',
            'url': (
                '/web/binary/saveas?model=ir.attachment&field=datas'
                '&filename_field=datas_fname&id=%s' % attachment.id
            ),
            'target': 'self',
        }

    @api.multi
    def _build_xlsx_attachment(self, lots):
        """Builds the workbook, saves it as an ir.attachment and returns
        that attachment record."""
        self.ensure_one()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Serial Numbers'

        headers = [
            'Customer',
            'Project',
            'Vendor',
            'Start Date',
            'End Date',
            'Part Number',
            'Serial Number',
        ]

        # headers = [
        #     'Part Number',
        #     'Customer',
        #     'Vendor',
        #     'Project',
        #     'Start Date',
        #     'End Date',
        #     'Serial Number',
        # ]

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4',
                                   fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')
        thin_side = Side(style='thin', color='B7B7B7')
        cell_border = Border(left=thin_side, right=thin_side,
                              top=thin_side, bottom=thin_side)

        # -- Header row --
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = cell_border
        ws.row_dimensions[1].height = 20

        # -- Data rows --
        row_idx = 2
        for lot in lots:
            part_number = lot.product_id.default_code or lot.product_id.name or ''
            customer = ''
            if lot.customer_id.is_company:
                customer = lot.customer_id.name
            else:
                customer = lot.customer_id.parent_id.name
            customer = customer
            vendor = lot.product_id.od_pdt_brand_id.name or ''
            project = lot.analytic_account_id.name or ''
            start_date = self._format_date(lot.use_date)
            end_date = self._format_date(lot.life_date)
            serial_number = lot.name or ''

            row_values = [
                customer, project, vendor, start_date,
                end_date, part_number, serial_number,
            ]

            # row_values = [
            #     part_number, customer, vendor, project,
            #     start_date, end_date, serial_number,
            # ]
            for col_idx, value in enumerate(row_values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = cell_border
                cell.alignment = Alignment(horizontal='left', vertical='center')
            row_idx += 1

        # -- Column widths --
        col_widths = [38, 40, 18, 14, 14, 20, 22]
        for idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        ws.freeze_panes = 'A2'

        # -- Save to memory and create attachment --
        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        file_data = base64.b64encode(file_stream.getvalue())

        filename = 'Product Subscription Expiration Report.xlsx'

        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'datas': file_data,
            'datas_fname': filename,
            # Explicit mimetype avoids any auto-detection ambiguity and
            # tells the browser exactly what kind of file it's getting.
            'mimetype': (
                'application/vnd.openxmlformats-officedocument'
                '.spreadsheetml.sheet'
            ),
            # Deliberately NOT linked to the transient wizard record
            # (res_model/res_id left empty): wizard rows are periodically
            # vacuumed by Odoo, which would eventually break the download
            # link / trip up access rules on the attachment.
            'res_model': False,
            'res_id': False,
            'type': 'binary',
        })
        return attachment

    @staticmethod
    def _format_date(value):
        """stock.production.lot date fields are stored as strings
        ('YYYY-MM-DD' or 'YYYY-MM-DD HH:MM:SS'). Return just the date
        part, or an empty string if not set."""
        if not value:
            return ''
        return value.split(' ')[0]
